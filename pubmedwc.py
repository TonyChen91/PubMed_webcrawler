#抓取網頁原始碼(html)
import bs4
import urllib.request as req
from openpyxl import Workbook, load_workbook
xlsxpath="C:/Users/tony/Desktop/實驗室pdf資料/PubMed爬蟲.xlsx" #檔案儲存位置
HowManyPages=5 #搜尋幾頁
pmcbaseurl="https://www.ncbi.nlm.nih.gov/pmc/articles/pmid"
pubmedurl="https://pubmed.ncbi.nlm.nih.gov/?term=%22alzheimer%22%20%22precision%22&filter=simsearch1.fha&page=" #搜尋網址
#模擬使用者(建立header資訊)
page=1
excelp = 1
def getData(url):
    global excelp
    request = req.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Mobile Safari/537.36"
    })
    #
    with req.urlopen(request) as response:
        data = response.read().decode("utf-8")
    #print(data)
    #解析資料
    root = bs4.BeautifulSoup(data, "html.parser")
    freearticles = root.find_all("span", class_="free-resources spaced-citation-item citation-part", string="Free PMC article.")
    for freearticle in freearticles:
        title = freearticle.find_parent("div", class_="docsum-citation full-citation").find_previous_sibling("a", class_="docsum-title")
        ws["A"+str(excelp)].value = title.text.strip()
        hyperlink= pmcbaseurl+title["href"]
        ws["B"+str(excelp)].value = '=HYPERLINK("%s", "LINK")' % hyperlink
        #print(title.text.strip())
        #print(pmcbaseurl+title["href"])
        excelp+=1
    return "&page="+str(page)
wb = load_workbook(xlsxpath)
wb.create_sheet("PubMed爬蟲共"+str(HowManyPages)+"頁")
ws = wb["PubMed爬蟲共"+str(HowManyPages)+"頁"]
while page <= HowManyPages:
    getData(pubmedurl+str(page))
    page+=1
wb.save(xlsxpath)
#    if GG.a != None:
#        print(GG.a.string)
#擴增上下頁##
#nextLink = root.find("a", string="‹ 上頁")  # 找a又有‹ 上頁的文字
#return nextLink["href"]  # print(nextLink)中href的屬性
